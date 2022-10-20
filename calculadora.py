import pandas as pd
import openpyxl
#from openpyxl import load_workbook, Workbook
import math
import funcao_calculadora
import os


def calculadora_carbono():
	funcao_calculadora.mensagem_inicial()
	
	planilha = openpyxl.load_workbook(filename = "planilha_medidas.xlsm")
	sheet_range = planilha['Planilha1']
	#sh2 = planilha['resultado']
	#tempo_total = transformar_anos(sheet_range)
	divisores = 1000000000
	valor_arvore = 0.541
	carbono_movel = funcao_calculadora.calculo_combustao_movel(sheet_range)

	carbono_aerea = funcao_calculadora.calculo_combustao_aerea(sheet_range)

	carbono_eletrica = funcao_calculadora.energia_eletrica(sheet_range)

	carbono_estacionaria = funcao_calculadora.calculo_combustao_estacionaria(sheet_range)

	carbono_final_mes = carbono_movel + carbono_eletrica + carbono_estacionaria

	carbono_final_ano = 12*(carbono_final_mes) + carbono_aerea
	carbono_final_mes += carbono_aerea

	total_arvores = math.ceil(carbono_final_mes/0.541)


	carbono_final_ano = round(carbono_final_ano,2)
	carbono_final_mes =  round(carbono_final_mes,2)

	#carbono_final = carbono_final#/ divisores

	#print(f'{carbono_final_mes:.3f} tCO2')
	#print(f'O total de árvores que você deve plantar é {total_arvores}')

	wb = openpyxl.load_workbook(filename = "resultado.xlsm")
	sh1 = wb['resultado']

	celula_quantidade_arvores = 'F6'
	celula_quantidade_arvores_mes = 'F7'
	celula_quantidade_carbono_ano = 'F8'

	escrever_celula(total_arvores,sh1,celula_quantidade_arvores)
	escrever_celula(carbono_final_mes,sh1,celula_quantidade_arvores_mes)
	escrever_celula(carbono_final_ano,sh1,celula_quantidade_carbono_ano)

	wb.save("resultado.xlsm")

	os.startfile("C:\\Users\\alexa\\Desktop\\APS-UNIP\\resultado.xlsm")

def escrever_celula(total_arvores,sh1, celula):
	sh1[celula].value = total_arvores



if (__name__ == '__main__'):
	calculadora_carbono()
